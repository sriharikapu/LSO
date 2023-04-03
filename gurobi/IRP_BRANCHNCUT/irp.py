import sys
import random
import math
from itertools import combinations
import gurobipy as gp
from gurobipy import GRB
import timeit
from openpyxl import Workbook
from openpyxl import load_workbook
from math import ceil
from math import floor
from math import sqrt
import numpy as np
import itertools
import networkx as nx


t0= timeit.default_timer()

wb = load_workbook("IRPData.xlsx")

##wtest = wb['TestData']
##wtest = wb['TestData1']
##wtest = wb['TestData2']
wtest = wb['TestData3']

nN = wtest['B1'].value # number of suppliers
N = [n+1 for n in range(nN)]
N0 = [0]+N
nT = wtest['H1'].value # number of periods
T = [t+1 for t in range(nT)]
T0 = [0]+T
S = []

## Defining edges for nodes and subsets:
Eall = [(i,j) for i in N0 for j in N0 if i<j] ## set of all edges
E = {} #edges in a subset
Del = {} #edges incident on a subset or a node, such that one node is within the subset and other outside
def Diff(li1, li2):
          return list(set(li1) - set(li2)) + list(set(li2) - set(li1))
for su in S:
   su1 = Diff(N,su)
   E[su] = [(i,j) for i in su for j in su if i<j]
   Del[su] = [(i,j) for i in su for j in su1 if i<j]+[(i,j) for i in su1 for j in su if i<j]
for i in N0:
   Del[i] = [(i,j) for j in N0 if i<j]+[(j,i) for j in N0 if j<i]

H = {} # inventory holding cost at customer i
C = {} # transportation cost i-j-k
D = {} # demand i-t
I = {} # initial inventory at customer i
L = {} # maximum inventory level at customer i
B = {} # shortage cost at customer i
Q = wtest.cell(row=3+nN+4,column=2).value
for i in N:
   H[i] = wtest.cell(row=i+3,column=2).value
   I[i]= wtest.cell(row=i+3,column=nT+nN+7).value
   L[i]= wtest.cell(row=i+3,column=nT+nN+8).value
   B[i]= wtest.cell(row=i+3,column=nT+nN+9).value
   for t in T:
       D[i,t] = wtest.cell(row=i+3,column=t+3).value
       for j in N0:
          C[i,j] = wtest.cell(row=i+3,column=j+nT+5).value\
                    *wtest.cell(row=nN+7,column=5).value\
                    /wtest.cell(row=3+nN+4,column=3).value
          C[j,i] = C[i,j]

# Callback - use lazy constraints to eliminate sub-tours
def subtourelim(model, where):
  if where == GRB.Callback.MIPSOL:
      #Z = model.cbGetSolution(model._z)
      X = model.cbGetSolution(model._x)
      X0 = model.cbGetSolution(model._x0)
      Qq = model.cbGetSolution(model._q)
      for t in T:
         edges = [(i,j) for i in N for j in N if i<j and X[i,j,t]>0.5]
         edges += [(0,i) for i in N if X0[0,i,t]>0.5]
         STset = subtourset(edges)
         if len(STset)>0:
             for su in STset:
                for e in su:
                   model.cbLazy((gp.quicksum(model._x[i,j,t] for i in su for j in su if i<j and i>0))\
                             <= gp.quicksum(model._z[i,t] for i in su if i>0)-model._z[e,t])
def subtourset(edges):
  G = nx.Graph()
  G.add_edges_from(edges)
  lo = list(nx.chain_decomposition(G, root=None))
  STset = []
  if len(lo)>0:
      for i in range(len(lo)):
         g1 = nx.Graph()
         g1.add_edges_from(lo[i])
         nds = list(nx.nodes(g1))
         STset.append(nds)
  return STset

#### model building starts:
m = gp.Model()
m.Params.TIME_LIMIT = 3600
## inventory level variable
sset = [(i,t) for i in N for t in T0]
s_ub = [L[i] for i in N for t in T0]
s = m.addVars(sset,lb=0,ub=s_ub, name="sit")
## shortage variables
bset = [(i,t) for i in N for t in T]
b_ub = [D[i,t] for i in N for t in T]
b = m.addVars(bset,lb=0,ub=b_ub, name="bit")
## routing variables x
xset = [(i,j,t) for i in N for j in N for t in T if i<j]
x = m.addVars(xset,vtype=GRB.BINARY,name="xijt")
x0set = [(0,i,t) for i in N for t in T]
x0 = m.addVars(x0set,vtype=GRB.INTEGER,lb=0,ub=2,name="x0it")
## Quantity delivery variables\
qset = [(i,t) for i in N for t in T]
q_ub = [Q for i in N for t in T]
q = m.addVars(qset,lb=0,ub=q_ub, name="qit")
## visit variables
zset = [(i,t) for i in N0 for t in T]
z = m.addVars(zset,vtype=GRB.BINARY,name="zit")

## Objective function:
m.setObjective(gp.quicksum(gp.quicksum(H[i]*s[i,t] for i in N)+gp.quicksum(B[i]*b[i,t] for i in N) +\
                    gp.quicksum(C[(key[0],key[1])]*x[key] for key in x.keys()) + \
                    gp.quicksum(C[(key[0],key[1])]*x0[key] for key in x0.keys()) for t in T))
for i in N:
   #c1: initial inventory at customer
   m.addConstr(s[i,0]==I[i])
   for t in T:
       #c2: inventory balance at each customer
       m.addConstr(s[i,t]-b[i,t] == s[i,t-1] + q[i,t] - D[i,t])
       m.addConstr(s[i,t-1]+q[i,t]<=L[i])
       ## no loading at a customer unless visited:
       m.addConstr(q[i,t] <= min(Q,L[i],sum(D[i,t] for w in range(t,nT+1)))*z[i,t])
       ## not more than two edges incident on any customer node using a vehicle type
       m.addConstr(gp.quicksum(x[j,j1,t] for (j,j1) in Del[i] if j>0)+ x0[0,i,t] == 2*z[i,t])
for t in T:
   m.addConstr(gp.quicksum(q[i,t] for i in N)<=Q*z[0,t])
   ## not more than two edges incident on mfr using a vehicle type for each vehicle
   m.addConstr(gp.quicksum(x0[0,j,t] for j in N) == 2*z[0,t])

t1= timeit.default_timer()

#### SOLN part Begins
# Optimize model
m._s = s
m._b = b
m._x = x
m._x0 = x0
m._q = q
m._z = z

m.Params.lazyConstraints = 1
m.optimize(subtourelim)
t2= timeit.default_timer()
Obj = m.objVal
print('Obj:',Obj,'--Time:',t2-t1)
Ss = m.getAttr('x', s)
Bs = m.getAttr('x', b)
X = m.getAttr('x', x)
X0 = m.getAttr('x', x0)
Qq = m.getAttr('x', q)
Z = m.getAttr('x', z)
### Print out values of decision variables:
print('Printing out values: ')
ObjCal = sum(sum(H[i]*Ss[i,t] for i in N) + sum(C[(key[0],key[1])]*X[key] for key in X.keys())\
           +sum(C[(key[0],key[1])]*X0[key] for key in X0.keys()) for t in T)
print("Objective w/0 shortage = ", ObjCal)
for t in T:
   print('\n\n+++++++++++++++++++t = ',t)
   print('Node visiting times Z: ')
   print('i - Zval')
   for i in N0:
      if Z[i,t]>0.2: print(i,' - ',round(Z[i,t],0))
   print('Quantity loaded q: ')
   print('i - Qq')
   for i in N:
      print(i,' - ',round(Qq[i,t],1))
   print('Routing variables')
   print('i - j - Xval')
   for i in N:
      if X0[0,i,t]>0.2: print(i,' - ',0,' - ',round(X0[0,i,t],0))
      for j in N:
          if i<j:
              if X[i,j,t]>0.2: print(i,' - ',j,' - ',round(X[i,j,t],0))
   print('Inventory / shortage levels at the end of period: ')
   print('i - sVal - bVal')
   for i in N:
      print(i,' - ',round(Ss[i,t],1),' - ',round(Bs[i,t],1))
